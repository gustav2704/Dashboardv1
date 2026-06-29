import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app import db
from app.catalog import import_catalog
from app.catalog_export import _build_records, export_catalog
from app.main import dashboard_data
from app.mapping import ensure_mt5_strategies


def test_excel_enriches_existing_mt5_strategy(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "enrichment.db")
    db.init_db()
    now = db.utcnow()
    with db.session() as conn:
        terminal_id = conn.execute(
            """INSERT INTO terminals(
                 name,data_dir,account_login,status,created_at
               ) VALUES(?,?,?,?,?)""",
            ("Fixture", str(tmp_path / "terminal"), "123", "connected", now),
        ).lastrowid
        conn.execute(
            """INSERT INTO deals(
                 terminal_id,ticket,position_id,time_msc,symbol,deal_type,entry_type,
                 volume,price,profit,commission,swap,magic,comment,raw_json
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                terminal_id,
                1,
                1,
                1000,
                "XAUUSD.cyr",
                "BUY",
                "IN",
                0.1,
                2300,
                0,
                0,
                0,
                77,
                "Exact MT5 name",
                json.dumps({}),
            ),
        )
    assert ensure_mt5_strategies()["created"] == 1

    source = tmp_path / "catalog.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "symbol",
            "SQX original name",
            "mql5 bot name (alternative)",
            "demo account number",
            "SQN",
        ]
    )
    sheet.append(["XAU", "SQX canonical name", "Exact MT5 name", "123", 2.1])
    workbook.save(source)

    import_catalog(source)

    with db.session() as conn:
        assert conn.execute("SELECT COUNT(*) FROM strategies").fetchone()[0] == 1
        strategy = conn.execute(
            "SELECT sqx_name,mql5_name,origin FROM strategies"
        ).fetchone()
        assert tuple(strategy) == (
            "SQX canonical name",
            "Exact MT5 name",
            "mt5+excel",
        )
        assert conn.execute("SELECT COUNT(*) FROM mappings").fetchone()[0] == 1


def test_export_preserves_source_and_builds_multisource_workbook(tmp_path, monkeypatch):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "export.db")
    db.init_db()
    source = Path(__file__).resolve().parents[3] / "EA_track" / "Track_v1.xlsx"
    original_bytes = source.read_bytes()
    import_catalog(source)
    strategies = dashboard_data()["strategies"]
    with db.session() as conn:
        catalog_json = {
            row["id"]: row["catalog_json"]
            for row in conn.execute("SELECT id,catalog_json FROM strategies")
        }
    export_rows = [
        {**strategy, "catalog_json": catalog_json[strategy["id"]]}
        for strategy in strategies
    ]
    destination = tmp_path / "updated.xlsx"

    export_catalog(source, destination, export_rows)

    assert source.read_bytes() == original_bytes
    exported = load_workbook(destination, read_only=False, data_only=False)
    assert "SQX_strategy list" in exported.sheetnames
    assert exported.sheetnames[:6] == [
        "Ranking",
        "Comparativo",
        "SQX",
        "Backtest MT5",
        "Live MT5",
        "Metodología",
    ]
    assert "Dashboard MT5" not in exported.sheetnames
    ranking = exported["Ranking"]
    comparison = exported["Comparativo"]
    assert ranking.max_row == len(strategies) + 9
    assert comparison.max_row == len(strategies) + 5
    assert ranking.freeze_panes == "D10"
    assert comparison.freeze_panes == "F6"
    comparison_headers = {
        cell.value: cell.column for cell in comparison[5] if cell.value
    }
    assert "Score promesa" in comparison_headers
    assert "Retención PF live/OOS" in comparison_headers
    assert "Datos faltantes" in comparison_headers
    score_formula = comparison.cell(6, comparison_headers["Score promesa"]).value
    retention_formula = comparison.cell(
        6, comparison_headers["Retención PF live/OOS"]
    ).value
    assert score_formula.startswith("=IF(")
    assert "IFERROR" in retention_formula
    assert len(ranking._charts) == 1
    calculated = load_workbook(destination, read_only=False, data_only=True)
    assert calculated["Ranking"]["D10"].value
    assert calculated["Ranking"]["F10"].value == 0
    assert calculated["Comparativo"]["J6"].value is not None


def test_export_uses_latest_completed_backtest_and_keeps_latest_failure_status(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "backtest-export.db")
    db.init_db()
    now = db.utcnow()
    with db.session() as conn:
        strategy_id = conn.execute(
            """INSERT INTO strategies(symbol,sqx_name,account_login,origin,created_at)
               VALUES(?,?,?,?,?)""",
            ("XAU", "Backtest export", "123", "sqx", now),
        ).lastrowid
        older = conn.execute(
            """INSERT INTO backtest_runs(
                 strategy_id,broker,expert_path,expert_hash,symbol,timeframe,
                 from_date,to_date,deposit,currency,leverage,model,config_source,
                 status,requested_at,finished_at
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                strategy_id, "FPM", "old.ex5", "old", "XAUUSD", "H1",
                "2020-01-01", "2024-01-01", 100000, "USD", "1:100", 1,
                "test", "completed", "2026-01-01T00:00:00+00:00",
                "2026-01-02T00:00:00+00:00",
            ),
        ).lastrowid
        newer = conn.execute(
            """INSERT INTO backtest_runs(
                 strategy_id,broker,expert_path,expert_hash,symbol,timeframe,
                 from_date,to_date,deposit,currency,leverage,model,config_source,
                 status,requested_at,finished_at
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                strategy_id, "FPM", "new.ex5", "new", "XAUUSD", "H1",
                "2021-01-01", "2025-01-01", 100000, "USD", "1:100", 1,
                "test", "completed", "2026-02-01T00:00:00+00:00",
                "2026-02-02T00:00:00+00:00",
            ),
        ).lastrowid
        conn.executemany(
            """INSERT INTO backtest_metrics(
                 run_id,metrics_json,raw_metrics_json,parsed_at
               ) VALUES(?,?,?,?)""",
            [
                (older, json.dumps({"profit_factor": 1.1, "trades": 100}), "{}", now),
                (newer, json.dumps({"profit_factor": 1.7, "trades": 200}), "{}", now),
            ],
        )
        conn.execute(
            """INSERT INTO backtest_runs(
                 strategy_id,broker,expert_path,expert_hash,symbol,timeframe,
                 from_date,to_date,deposit,currency,leverage,model,config_source,
                 status,requested_at,finished_at,error
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                strategy_id, "FPM", "failed.ex5", "failed", "XAUUSD", "H1",
                "2021-01-01", "2025-01-01", 100000, "USD", "1:100", 1,
                "test", "failed", "2026-03-01T00:00:00+00:00",
                "2026-03-02T00:00:00+00:00", "fixture failure",
            ),
        )
    strategy = {
        "id": strategy_id,
        "symbol": "XAU",
        "sqx_name": "Backtest export",
        "mql5_name": "",
        "account_login": "123",
        "origin": "sqx",
        "state": "unlinked",
        "link_state": "sqx_only",
        "sqx": None,
        "metrics": {"trades": 0},
        "risk_guard": {"status": "gray"},
        "baselines": [],
    }

    record = _build_records([strategy])[0]

    assert record["completed"]["id"] == newer
    assert record["bt"]["profit_factor"] == 1.7
    assert record["latest_run"]["status"] == "failed"
    assert "Backtest MT5" not in record["missing"]


def test_red_risk_overrides_ranking_category_and_missing_sources_are_explicit(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(db, "DB_PATH", tmp_path / "risk-export.db")
    db.init_db()
    strategy = {
        "id": 1,
        "symbol": "NAQ",
        "sqx_name": "Risk fixture",
        "mql5_name": "",
        "account_login": "",
        "origin": "mt5",
        "state": "active",
        "link_state": "mt5_only",
        "sqx": None,
        "metrics": {
            "trades": 10,
            "profit_factor": 1.2,
            "return_dd": 0.5,
            "expectancy": 5,
            "avg_loss": -20,
            "trades_per_month": 2,
            "max_drawdown": 100,
            "max_consecutive_losses": 3,
            "current_consecutive_losses": 2,
        },
        "risk_guard": {"status": "red"},
        "baselines": [],
    }

    record = _build_records([strategy])[0]

    assert record["category"] == "Revisar/Pausar"
    assert "Edge: N/D" in record["missing"]
    assert "EGT: N/D" in record["missing"]
    assert "SQX OOS: N/D" in record["missing"]
