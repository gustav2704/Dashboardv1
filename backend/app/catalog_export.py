from __future__ import annotations

import json
import os
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .db import session


GENERATED_SHEETS = [
    "Ranking",
    "Comparativo",
    "SQX",
    "Backtest MT5",
    "Live MT5",
    "Metodología",
    "Dashboard MT5",
]
SOURCE_COLORS = {
    "identity": "334155",
    "ranking": "475569",
    "sqx": "0F766E",
    "backtest": "2563EB",
    "live": "B45309",
    "risk": "B91C1C",
    "comparison": "0E7490",
}
COMPONENT_WEIGHTS = {"sqx": 35.0, "backtest": 25.0, "live": 25.0, "risk": 15.0}
SQX_SUBWEIGHTS = {"edge": 0.30, "egt": 0.20, "oos": 0.30, "decay": 0.20}
SHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _json(value: Any, default: Any = None) -> Any:
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value or "")
    except (TypeError, json.JSONDecodeError):
        return {} if default is None else default


def _fold(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").casefold())


def _number(source: dict[str, Any] | None, *keys: str) -> float | None:
    if not source:
        return None
    folded = {_fold(key): value for key, value in source.items()}
    for key in keys:
        value = folded.get(_fold(key))
        if isinstance(value, dict):
            value = value.get("amount", value.get("percent"))
        if value in (None, ""):
            continue
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return None


def _nested_number(source: dict[str, Any] | None, key: str, child: str) -> float | None:
    value = (source or {}).get(key)
    if not isinstance(value, dict):
        return None
    try:
        return float(value.get(child))
    except (TypeError, ValueError):
        return None


def _clamp(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, value))


def _linear(value: float | None, low: float, high: float, higher: bool = True) -> float | None:
    if value is None or high == low:
        return None
    score = (value - low) / (high - low) * 100
    return _clamp(score if higher else 100 - score)


def _average(values: Iterable[float | None]) -> float | None:
    present = [float(value) for value in values if value is not None]
    return sum(present) / len(present) if present else None


def _weighted_available(items: Iterable[tuple[float | None, float]]) -> tuple[float | None, float]:
    present = [(float(score), weight) for score, weight in items if score is not None]
    available_weight = sum(weight for _, weight in present)
    if not available_weight:
        return None, 0.0
    return (
        sum(score * weight for score, weight in present) / available_weight,
        available_weight,
    )


def _closeness(left: float | None, right: float | None) -> float | None:
    if left is None or right is None or left <= 0 or right <= 0:
        return None
    return _clamp(min(left / right, right / left) * 100)


def _win_rate(value: float | None) -> float | None:
    if value is None:
        return None
    return value / 100 if value > 1 else value


def _risk_headroom(actual: float | None, limit: float | None) -> float | None:
    if actual is None or limit is None or limit <= 0:
        return None
    ratio = actual / limit
    if ratio <= 0.5:
        return 100.0
    if ratio >= 1:
        return 0.0
    return (1 - ratio) / 0.5 * 100


def _joined(values: Iterable[Any]) -> str:
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in result:
            result.append(text)
    return ", ".join(result)


def _baseline_map(strategy: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for snapshot in strategy.get("baselines", []):
        key = (
            str(snapshot.get("source") or "").casefold(),
            str(snapshot.get("sample_type") or "").casefold(),
        )
        result.setdefault(key, snapshot.get("metrics", {}))
    return result


def _latest_related() -> dict[str, dict[int, Any]]:
    mappings: dict[int, list[dict[str, Any]]] = {}
    analytics: dict[int, dict[str, Any]] = {}
    latest_runs: dict[int, dict[str, Any]] = {}
    completed_runs: dict[int, dict[str, Any]] = {}
    with session() as conn:
        for row in conn.execute(
            """SELECT m.*,t.name terminal_name,t.server,t.last_sync
               FROM mappings m JOIN terminals t ON t.id=m.terminal_id
               WHERE m.confirmed=1 ORDER BY m.strategy_id,m.id"""
        ):
            item = dict(row)
            mappings.setdefault(int(item["strategy_id"]), []).append(item)
        for row in conn.execute(
            """SELECT strategy_id,analytics_json,synced_at
               FROM sqx_analytics_snapshots ORDER BY synced_at DESC,id DESC"""
        ):
            strategy_id = int(row["strategy_id"])
            analytics.setdefault(
                strategy_id,
                {**_json(row["analytics_json"]), "synced_at": row["synced_at"]},
            )
        for row in conn.execute(
            "SELECT * FROM backtest_runs ORDER BY requested_at DESC,id DESC"
        ):
            item = dict(row)
            latest_runs.setdefault(int(item["strategy_id"]), item)
        for row in conn.execute(
            """SELECT r.*,m.metrics_json,m.raw_metrics_json,m.parsed_at
               FROM backtest_runs r JOIN backtest_metrics m ON m.run_id=r.id
               WHERE r.status='completed'
               ORDER BY COALESCE(r.finished_at,r.requested_at) DESC,r.id DESC"""
        ):
            item = dict(row)
            item["metrics"] = _json(item.pop("metrics_json"))
            item["raw_metrics"] = _json(item.pop("raw_metrics_json"))
            item["config_snapshot"] = _json(item.get("config_snapshot_json"))
            completed_runs.setdefault(int(item["strategy_id"]), item)
    return {
        "mappings": mappings,
        "analytics": analytics,
        "latest_runs": latest_runs,
        "completed_runs": completed_runs,
    }


def _sqx_scores(
    analytics: dict[str, Any],
    oos: dict[str, Any],
) -> tuple[float | None, float, dict[str, float | None]]:
    edge = analytics.get("edge", {})
    egt = analytics.get("egt", {})
    edge_score = _number(edge, "score") if edge.get("available") else None
    egt_total = _number(egt, "total") if egt.get("available") else None
    egt_score = _clamp(egt_total * 10) if egt_total is not None else None
    oos_quality = _average(
        [
            _linear(_number(oos, "ProfitFactor"), 1, 2),
            _linear(_number(oos, "ReturnDDRatio"), 0, 3),
            _linear(_number(oos, "SQN"), 0, 3),
            _linear(_number(oos, "SharpeRatio"), 0, 2),
        ]
    )
    decay = analytics.get("decay_pct", {})
    decay_score = _average(
        [
            None
            if _number(decay, key) is None
            else _clamp(100 - max(float(_number(decay, key) or 0), 0))
            for key in (
                "ProfitFactor",
                "NetProfit",
                "ReturnDDRatio",
                "WinningPct",
                "SharpeRatio",
                "Stability",
                "PctDrawdown",
            )
        ]
    )
    score, coverage = _weighted_available(
        [
            (edge_score, SQX_SUBWEIGHTS["edge"]),
            (egt_score, SQX_SUBWEIGHTS["egt"]),
            (oos_quality, SQX_SUBWEIGHTS["oos"]),
            (decay_score, SQX_SUBWEIGHTS["decay"]),
        ]
    )
    return score, coverage, {
        "edge_score": edge_score,
        "egt_score": egt_score,
        "oos_quality": oos_quality,
        "decay_score": decay_score,
    }


def _backtest_scores(
    metrics: dict[str, Any],
    full: dict[str, Any],
) -> tuple[float | None, float, dict[str, float | None]]:
    if not metrics:
        return None, 0.0, {"quality": None, "agreement": None}
    dd_pct = (
        _nested_number(metrics, "equity_drawdown_max", "percent")
        or _nested_number(metrics, "equity_drawdown_relative", "percent")
    )
    quality, quality_coverage = _weighted_available(
        [
            (_linear(_number(metrics, "profit_factor"), 1, 2), 0.25),
            (_linear(_number(metrics, "recovery_factor"), 0, 5), 0.20),
            (_linear(_number(metrics, "sharpe_ratio"), 0, 3), 0.15),
            (_linear(dd_pct, 0, 20, higher=False), 0.20),
            (_linear(_number(metrics, "trades"), 0, 300), 0.20),
        ]
    )
    bt_win = _win_rate(_number(metrics, "win_rate"))
    sqx_win = _win_rate(_number(full, "WinningPct"))
    win_agreement = (
        _clamp(100 - abs(bt_win - sqx_win) * 500)
        if bt_win is not None and sqx_win is not None
        else None
    )
    agreement = _average(
        [
            _closeness(_number(metrics, "profit_factor"), _number(full, "ProfitFactor")),
            _closeness(dd_pct, _number(full, "PctDrawdown")),
            _closeness(_number(metrics, "trades_per_month"), _number(full, "AvgTradesPerMonth")),
            win_agreement,
        ]
    )
    score, outer_coverage = _weighted_available(
        [(quality, 0.50), (agreement, 0.50)]
    )
    coverage = (
        0.50 * quality_coverage + (0.50 if agreement is not None else 0.0)
    )
    return score, min(coverage, outer_coverage), {
        "quality": quality,
        "agreement": agreement,
        "dd_pct": dd_pct,
    }


def _live_scores(
    live: dict[str, Any],
    oos: dict[str, Any],
) -> tuple[float | None, float, dict[str, float | None]]:
    trades = _number(live, "trades") or 0
    if trades <= 0:
        return None, 0.0, {"expectancy_ratio": None}
    avg_loss = abs(_number(live, "avg_loss") or 0)
    expectancy = _number(live, "expectancy")
    expectancy_ratio = expectancy / avg_loss if expectancy is not None and avg_loss else None
    score, coverage = _weighted_available(
        [
            (_linear(_number(live, "profit_factor"), 1, 2), 0.20),
            (_linear(_number(live, "return_dd"), 0, 3), 0.20),
            (_linear(expectancy_ratio, 0, 0.5), 0.15),
            (_linear(trades, 0, 50), 0.15),
            (
                None
                if _number(oos, "ProfitFactor") in (None, 0)
                else _clamp(
                    (_number(live, "profit_factor") or 0)
                    / float(_number(oos, "ProfitFactor") or 1)
                    * 100
                ),
                0.15,
            ),
            (
                _closeness(
                    _number(live, "trades_per_month"),
                    _number(oos, "AvgTradesPerMonth"),
                ),
                0.15,
            ),
        ]
    )
    return score, coverage, {"expectancy_ratio": expectancy_ratio}


def _risk_scores(
    live: dict[str, Any],
    oos: dict[str, Any],
) -> tuple[float | None, float]:
    return _weighted_available(
        [
            (
                _risk_headroom(
                    _number(live, "max_drawdown"),
                    _number(oos, "Drawdown", "MaxDD", "MaxDrawdown"),
                ),
                0.50,
            ),
            (
                _risk_headroom(
                    _number(live, "max_consecutive_losses"),
                    _number(oos, "MaxConsecLoss", "MaxConsecutiveLosses"),
                ),
                0.30,
            ),
            (
                _risk_headroom(
                    _number(live, "current_consecutive_losses"),
                    _number(oos, "MaxConsecLoss", "MaxConsecutiveLosses"),
                ),
                0.20,
            ),
        ]
    )


def _build_records(strategies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    related = _latest_related()
    records: list[dict[str, Any]] = []
    for strategy in strategies:
        strategy_id = int(strategy["id"])
        mappings = related["mappings"].get(strategy_id, [])
        analytics = related["analytics"].get(strategy_id, {})
        baselines = _baseline_map(strategy)
        full = baselines.get(("sqx", "full"), {})
        inside = baselines.get(("sqx", "is"), {})
        oos = baselines.get(("sqx", "oos"), {})
        completed = related["completed_runs"].get(strategy_id, {})
        latest_run = related["latest_runs"].get(strategy_id, {})
        bt_metrics = completed.get("metrics", {})
        live = strategy.get("metrics", {})
        sqx_score, sqx_coverage, sqx_parts = _sqx_scores(analytics, oos)
        bt_score, bt_coverage, bt_parts = _backtest_scores(bt_metrics, full)
        live_score, live_coverage, live_parts = _live_scores(live, oos)
        risk_score, risk_coverage = _risk_scores(live, oos)
        component_scores = {
            "sqx": sqx_score,
            "backtest": bt_score,
            "live": live_score,
            "risk": risk_score,
        }
        component_coverage = {
            "sqx": sqx_coverage,
            "backtest": bt_coverage,
            "live": live_coverage,
            "risk": risk_coverage,
        }
        coverage = sum(
            COMPONENT_WEIGHTS[key] * component_coverage[key]
            for key in COMPONENT_WEIGHTS
        )
        quality = (
            sum(
                float(component_scores[key] or 0)
                * COMPONENT_WEIGHTS[key]
                * component_coverage[key]
                for key in COMPONENT_WEIGHTS
            )
            / coverage
            if coverage
            else None
        )
        promise = (
            quality * (0.5 + 0.5 * coverage / 100)
            if quality is not None
            else None
        )
        risk_status = str(strategy.get("risk_guard", {}).get("status") or "gray")
        if risk_status == "red":
            category = "Revisar/Pausar"
        elif coverage < 35:
            category = "Datos insuficientes"
        elif promise is not None and promise >= 75 and coverage >= 75:
            category = "Prioridad alta"
        elif promise is not None and promise >= 60 and coverage >= 50:
            category = "Prometedora"
        else:
            category = "Observar"

        edge = analytics.get("edge", {})
        egt = analytics.get("egt", {})
        missing: list[str] = []
        if not edge.get("available"):
            missing.append(f"Edge: {edge.get('reason') or 'N/D'}")
        if not egt.get("available"):
            missing.append(f"EGT: {egt.get('reason') or 'N/D'}")
        if not completed:
            missing.append(
                f"Backtest MT5: {latest_run.get('status') or 'N/D'}"
            )
        if (_number(live, "trades") or 0) <= 0:
            missing.append("Live: sin trades")
        if not oos:
            missing.append("SQX OOS: N/D")

        records.append(
            {
                "strategy": strategy,
                "mappings": mappings,
                "analytics": analytics,
                "full": full,
                "is": inside,
                "oos": oos,
                "completed": completed,
                "latest_run": latest_run,
                "bt": bt_metrics,
                "live": live,
                "scores": component_scores,
                "component_coverage": component_coverage,
                "coverage": coverage,
                "quality": quality,
                "promise": promise,
                "category": category,
                "risk_status": risk_status,
                "missing": "; ".join(missing),
                "sqx_parts": sqx_parts,
                "bt_parts": bt_parts,
                "live_parts": live_parts,
            }
        )
    return records


def _title(sheet: Any, title: str, subtitle: str, end_col: int) -> None:
    end = get_column_letter(min(end_col, 10))
    sheet.merge_cells(f"A1:{end}1")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor="102A43")
    sheet["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32
    sheet.merge_cells(f"A2:{end}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(color="52606D", italic=True, size=10)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22
    sheet.sheet_view.showGridLines = False


def _table(
    sheet: Any,
    headers: list[str],
    rows: list[list[Any]],
    start_row: int,
    table_name: str,
    color: str,
) -> dict[str, int]:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="CBD5E1"))
    sheet.row_dimensions[start_row].height = 34
    for values in rows:
        sheet.append(values)
    end_row = start_row + len(rows)
    end_col = len(headers)
    if rows:
        table = Table(
            displayName=table_name,
            ref=f"A{start_row}:{get_column_letter(end_col)}{end_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    for row in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=max(start_row + 1, end_row),
        max_col=end_col,
    ):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    return {header: index + 1 for index, header in enumerate(headers)}


def _widths(sheet: Any, headers: list[str], overrides: dict[str, float] | None = None) -> None:
    overrides = overrides or {}
    for index, header in enumerate(headers, start=1):
        width = overrides.get(header, max(11, min(22, len(header) + 2)))
        sheet.column_dimensions[get_column_letter(index)].width = width


def _format_columns(
    sheet: Any,
    indexes: dict[str, int],
    start_row: int,
    end_row: int,
    percent: Iterable[str] = (),
    money: Iterable[str] = (),
    decimals: Iterable[str] = (),
    integers: Iterable[str] = (),
    ratios: Iterable[str] = (),
) -> None:
    formats = [
        (percent, "0.0%"),
        (money, "#,##0.00"),
        (decimals, "0.00"),
        (integers, "#,##0"),
        (ratios, '0.00"x"'),
    ]
    for headers, number_format in formats:
        for header in headers:
            column = indexes.get(header)
            if not column:
                continue
            for row in range(start_row, end_row + 1):
                sheet.cell(row, column).number_format = number_format


def _status_text(payload: dict[str, Any], label: str) -> str:
    if payload.get("available"):
        return "Disponible"
    return f"N/D: {payload.get('reason') or label}"


def _cache_formula(
    cache: dict[str, dict[str, Any]],
    sheet_name: str,
    cell: Any,
    value: Any,
) -> None:
    cache.setdefault(sheet_name, {})[cell.coordinate] = value


def _write_formula_cache(
    destination: Path,
    cache: dict[str, dict[str, Any]],
) -> None:
    """Store calculated values beside formulas without removing the formulas."""
    fd, tmp_name = tempfile.mkstemp(
        prefix=destination.stem,
        suffix=".xlsx",
        dir=destination.parent,
    )
    os.close(fd)
    try:
        with zipfile.ZipFile(destination) as archive:
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            relations = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            targets = {
                node.attrib["Id"]: node.attrib["Target"]
                for node in relations.findall(f"{{{PKG_REL_NS}}}Relationship")
            }
            sheet_targets: dict[str, str] = {}
            for node in workbook.findall(f".//{{{SHEET_NS}}}sheet"):
                relation_id = node.attrib[f"{{{REL_NS}}}id"]
                target = targets[relation_id].lstrip("/")
                sheet_targets[node.attrib["name"]] = (
                    target if target.startswith("xl/") else f"xl/{target}"
                )
            with zipfile.ZipFile(tmp_name, "w", zipfile.ZIP_DEFLATED) as output:
                for item in archive.infolist():
                    payload = archive.read(item.filename)
                    sheet_name = next(
                        (
                            name
                            for name, target in sheet_targets.items()
                            if target == item.filename and name in cache
                        ),
                        None,
                    )
                    if sheet_name:
                        root = ET.fromstring(payload)
                        cells = {
                            node.attrib.get("r"): node
                            for node in root.findall(f".//{{{SHEET_NS}}}c")
                        }
                        for reference, value in cache[sheet_name].items():
                            cell = cells.get(reference)
                            if cell is None:
                                continue
                            value_node = cell.find(f"{{{SHEET_NS}}}v")
                            if value_node is None:
                                value_node = ET.SubElement(cell, f"{{{SHEET_NS}}}v")
                            if isinstance(value, str):
                                cell.set("t", "str")
                                value_node.text = value
                            elif value is None:
                                value_node.text = ""
                            else:
                                cell.attrib.pop("t", None)
                                value_node.text = str(float(value))
                        payload = ET.tostring(
                            root,
                            encoding="utf-8",
                            xml_declaration=True,
                        )
                    output.writestr(item, payload)
        os.replace(tmp_name, destination)
    finally:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)


def _sqx_sheet(workbook: Any, records: list[dict[str, Any]], generated_at: str) -> None:
    sheet = workbook["SQX"]
    headers = [
        "ID", "Estrategia", "Símbolo", "Proyecto", "Databank", "Filtro",
        "Timeframe", "Última sync", "Estado Edge", "Edge Score", "Edge Grade",
        "Pilar rentabilidad", "Pilar consistencia", "Pilar riesgo", "Pilar entrada",
        "XS", "Config Edge", "Estado EGT", "EGT total", "EGT grade", "EGT buy",
        "EGT sell", "N buy", "N sell", "Meses EGT", "Fuente EGT",
        "Full PF", "Full Return/DD", "Full SQN", "Full Sharpe", "Full Stability",
        "Full DD %", "Full trades/mes", "Full trades", "Full racha pérdida",
        "IS PF", "IS Return/DD", "IS SQN", "IS Sharpe", "IS Stability", "IS DD %",
        "IS trades/mes", "IS trades", "IS racha pérdida",
        "OOS PF", "OOS Return/DD", "OOS SQN", "OOS Sharpe", "OOS Stability",
        "OOS DD %", "OOS trades/mes", "OOS trades", "OOS racha pérdida",
        "Decay PF %", "Decay Net Profit %", "Decay Return/DD %", "Decay Win % %",
        "Decay Sharpe %", "Decay Stability %", "Aumento DD %", "Score SQX",
        "Cobertura SQX",
    ]
    rows: list[list[Any]] = []
    for record in records:
        strategy = record["strategy"]
        link = strategy.get("sqx") or {}
        analytics = record["analytics"]
        edge = analytics.get("edge", {})
        egt = analytics.get("egt", {})
        pillars = edge.get("pillars", {})
        decay = analytics.get("decay_pct", {})
        samples: list[Any] = []
        for sample in (record["full"], record["is"], record["oos"]):
            samples.extend(
                [
                    _number(sample, "ProfitFactor"),
                    _number(sample, "ReturnDDRatio"),
                    _number(sample, "SQN"),
                    _number(sample, "SharpeRatio"),
                    _number(sample, "Stability"),
                    None
                    if _number(sample, "PctDrawdown") is None
                    else _number(sample, "PctDrawdown") / 100,
                    _number(sample, "AvgTradesPerMonth"),
                    _number(sample, "NumberOfTrades"),
                    _number(sample, "MaxConsecLoss"),
                ]
            )
        rows.append(
            [
                strategy["id"], strategy["sqx_name"], strategy["symbol"],
                link.get("project"), link.get("databank"), link.get("filter_result"),
                link.get("timeframe"), link.get("last_synced_at"),
                _status_text(edge, "Edge"), _number(edge, "score"), edge.get("grade"),
                _number(pillars, "profitability"), _number(pillars, "consistency"),
                _number(pillars, "risk"), _number(pillars, "entry"),
                _number(edge, "xs_value"), edge.get("config_source"),
                _status_text(egt, "EGT"), _number(egt, "total"), egt.get("grade"),
                _number(egt, "buy"), _number(egt, "sell"), _number(egt, "n_buy"),
                _number(egt, "n_sell"), _number(egt, "months"),
                egt.get("history_source"), *samples,
                _number(decay, "ProfitFactor"), _number(decay, "NetProfit"),
                _number(decay, "ReturnDDRatio"), _number(decay, "WinningPct"),
                _number(decay, "SharpeRatio"), _number(decay, "Stability"),
                _number(decay, "PctDrawdown"), record["scores"]["sqx"],
                record["component_coverage"]["sqx"],
            ]
        )
    _title(sheet, "SQX | Robustez, Edge y EGT", f"Snapshot generado {generated_at}", len(headers))
    indexes = _table(sheet, headers, rows, 4, "SQXAnalyticsTable", SOURCE_COLORS["sqx"])
    sheet.freeze_panes = "D5"
    _widths(
        sheet,
        headers,
        {
            "Estrategia": 42,
            "Proyecto": 16,
            "Databank": 16,
            "Última sync": 24,
            "Estado Edge": 26,
            "Estado EGT": 30,
        },
    )
    _format_columns(
        sheet, indexes, 5, 4 + len(rows),
        percent=["Full DD %", "IS DD %", "OOS DD %", "Cobertura SQX"],
        decimals=[
            "Edge Score", "EGT total", "EGT buy", "EGT sell", "XS",
            "Full PF", "Full Return/DD", "Full SQN", "Full Sharpe", "Full Stability",
            "Full trades/mes", "IS PF", "IS Return/DD", "IS SQN", "IS Sharpe",
            "IS Stability", "IS trades/mes", "OOS PF", "OOS Return/DD", "OOS SQN",
            "OOS Sharpe", "OOS Stability", "OOS trades/mes", "Decay PF %",
            "Decay Net Profit %", "Decay Return/DD %", "Decay Win % %",
            "Decay Sharpe %", "Decay Stability %", "Aumento DD %", "Score SQX",
        ],
        integers=[
            "N buy", "N sell", "Meses EGT", "Full trades", "Full racha pérdida",
            "IS trades", "IS racha pérdida", "OOS trades", "OOS racha pérdida",
        ],
    )
    if rows:
        sheet.conditional_formatting.add(
            f"{get_column_letter(indexes['Edge Score'])}5:{get_column_letter(indexes['Edge Score'])}{4 + len(rows)}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                           mid_type="num", mid_value=60, mid_color="FFEB84",
                           end_type="num", end_value=100, end_color="63BE7B"),
        )


def _backtest_sheet(workbook: Any, records: list[dict[str, Any]], generated_at: str) -> None:
    sheet = workbook["Backtest MT5"]
    headers = [
        "ID", "Estrategia", "Símbolo", "Estado último intento", "Motivo / error",
        "Último completado", "Broker", "Timeframe", "Desde", "Hasta", "Modelo",
        "Spread", "Depósito", "Moneda", "Leverage", "Fuente config",
        "Calidad histórica", "Barras", "Ticks", "Trades", "Trades/mes",
        "Net profit", "Profit factor", "Expectancy", "Recovery factor", "Sharpe",
        "DD máximo", "DD máximo %", "Win rate", "Racha máxima pérdidas",
        "Avg win", "Avg loss", "Mejor trade", "Peor trade", "Long trades",
        "Long win rate", "Short trades", "Short win rate", "Score calidad",
        "Concordancia SQX", "Score backtest", "Cobertura backtest", "Reporte",
    ]
    rows: list[list[Any]] = []
    for record in records:
        strategy = record["strategy"]
        run = record["completed"]
        latest = record["latest_run"]
        metrics = record["bt"]
        rows.append(
            [
                strategy["id"], strategy["sqx_name"], strategy["symbol"],
                latest.get("status") or "N/D", latest.get("error") or (
                    "" if run else "N/D: sin backtest MT5 completado"
                ),
                run.get("finished_at"), run.get("broker"), run.get("timeframe"),
                run.get("from_date"), run.get("to_date"), run.get("model"),
                run.get("spread"), run.get("deposit"), run.get("currency"),
                run.get("leverage"), run.get("config_source"),
                None if _number(metrics, "history_quality") is None else _number(metrics, "history_quality") / 100,
                _number(metrics, "bars"), _number(metrics, "ticks"),
                _number(metrics, "trades"), _number(metrics, "trades_per_month"),
                _number(metrics, "net_profit"), _number(metrics, "profit_factor"),
                _number(metrics, "expectancy"), _number(metrics, "recovery_factor"),
                _number(metrics, "sharpe_ratio"), _number(metrics, "max_drawdown"),
                None if record["bt_parts"].get("dd_pct") is None else record["bt_parts"]["dd_pct"] / 100,
                _win_rate(_number(metrics, "win_rate")),
                _number(metrics, "max_consecutive_losses"), _number(metrics, "avg_win"),
                _number(metrics, "avg_loss"), _number(metrics, "best_trade"),
                _number(metrics, "worst_trade"), _number(metrics, "long_trades"),
                _win_rate(_number(metrics, "long_trades_win_rate")),
                _number(metrics, "short_trades"),
                _win_rate(_number(metrics, "short_trades_win_rate")),
                record["bt_parts"].get("quality"), record["bt_parts"].get("agreement"),
                record["scores"]["backtest"], record["component_coverage"]["backtest"],
                run.get("report_path"),
            ]
        )
    _title(sheet, "Backtest MT5 | Validación en broker", f"Último resultado completado por estrategia · {generated_at}", len(headers))
    indexes = _table(sheet, headers, rows, 4, "BacktestMT5Table", SOURCE_COLORS["backtest"])
    sheet.freeze_panes = "D5"
    _widths(
        sheet,
        headers,
        {
            "Estrategia": 42,
            "Estado último intento": 20,
            "Motivo / error": 34,
            "Último completado": 24,
            "Reporte": 42,
        },
    )
    _format_columns(
        sheet, indexes, 5, 4 + len(rows),
        percent=[
            "Calidad histórica", "DD máximo %", "Win rate", "Long win rate",
            "Short win rate", "Cobertura backtest",
        ],
        money=["Depósito", "Net profit", "Expectancy", "DD máximo", "Avg win", "Avg loss", "Mejor trade", "Peor trade"],
        decimals=[
            "Spread", "Trades/mes", "Profit factor", "Recovery factor", "Sharpe",
            "Score calidad", "Concordancia SQX", "Score backtest",
        ],
        integers=["Modelo", "Barras", "Ticks", "Trades", "Racha máxima pérdidas", "Long trades", "Short trades"],
    )
    for row in range(5, 5 + len(rows)):
        sheet.cell(row, indexes["Motivo / error"]).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        sheet.row_dimensions[row].height = 28


def _live_sheet(workbook: Any, records: list[dict[str, Any]], generated_at: str) -> None:
    sheet = workbook["Live MT5"]
    headers = [
        "ID", "Estrategia", "Símbolo catálogo", "Estado operativo", "Estado vínculo",
        "Cuenta MT5", "Terminal", "Servidor", "Símbolo broker", "Magic", "Comentario",
        "Última sync", "Trades", "Posiciones abiertas", "P/L neto", "P/L flotante",
        "Gross profit", "Gross loss", "Win rate", "Profit factor", "Expectancy",
        "Return / DD", "SQN", "Trades/mes", "Max drawdown", "Mejor trade",
        "Peor trade", "Avg win", "Avg loss", "Racha máx. ganancias",
        "Racha máx. pérdidas", "Racha actual pérdidas", "Comisiones", "Swaps",
        "Score live", "Cobertura live",
    ]
    rows: list[list[Any]] = []
    for record in records:
        strategy = record["strategy"]
        mappings = record["mappings"]
        live = record["live"]
        rows.append(
            [
                strategy["id"], strategy["sqx_name"], strategy["symbol"],
                strategy.get("state"), strategy.get("link_state"),
                _joined(mapping.get("account_login") for mapping in mappings)
                or strategy.get("account_login"),
                _joined(mapping.get("terminal_name") for mapping in mappings),
                _joined(mapping.get("server") for mapping in mappings),
                _joined(mapping.get("symbol") for mapping in mappings),
                _joined(mapping.get("magic") for mapping in mappings),
                _joined(mapping.get("comment_pattern") for mapping in mappings),
                max((str(mapping.get("last_sync") or "") for mapping in mappings), default=""),
                _number(live, "trades"), _number(live, "open_positions"),
                _number(live, "net_profit"), _number(live, "floating_profit"),
                _number(live, "gross_profit"), _number(live, "gross_loss"),
                _win_rate(_number(live, "win_rate")), _number(live, "profit_factor"),
                _number(live, "expectancy"), _number(live, "return_dd"),
                _number(live, "sqn"), _number(live, "trades_per_month"),
                _number(live, "max_drawdown"), _number(live, "best_trade"),
                _number(live, "worst_trade"), _number(live, "avg_win"),
                _number(live, "avg_loss"), _number(live, "max_consecutive_wins"),
                _number(live, "max_consecutive_losses"),
                _number(live, "current_consecutive_losses"),
                _number(live, "commissions"), _number(live, "swaps"),
                record["scores"]["live"], record["component_coverage"]["live"],
            ]
        )
    _title(sheet, "Live MT5 | Ejecución real en broker", f"Historial completo recibido del bridge · {generated_at}", len(headers))
    indexes = _table(sheet, headers, rows, 4, "LiveMT5Table", SOURCE_COLORS["live"])
    sheet.freeze_panes = "F5"
    _widths(
        sheet,
        headers,
        {
            "Estrategia": 42,
            "Estado operativo": 18,
            "Estado vínculo": 18,
            "Terminal": 18,
            "Servidor": 30,
            "Símbolo broker": 20,
            "Comentario": 38,
            "Última sync": 25,
        },
    )
    _format_columns(
        sheet, indexes, 5, 4 + len(rows),
        percent=["Win rate", "Cobertura live"],
        money=[
            "P/L neto", "P/L flotante", "Gross profit", "Gross loss", "Expectancy",
            "Max drawdown", "Mejor trade", "Peor trade", "Avg win", "Avg loss",
            "Comisiones", "Swaps",
        ],
        decimals=["Profit factor", "Return / DD", "SQN", "Trades/mes", "Score live"],
        integers=[
            "Trades", "Posiciones abiertas", "Racha máx. ganancias",
            "Racha máx. pérdidas", "Racha actual pérdidas",
        ],
    )


def _methodology_sheet(workbook: Any, records: list[dict[str, Any]], generated_at: str) -> None:
    sheet = workbook["Metodología"]
    _title(sheet, "Metodología | Ranking multifuente", f"Pesos y reglas auditables · {generated_at}", 9)
    sheet["A4"] = "Componente"
    sheet["B4"] = "Peso"
    for cell in sheet[4][:2]:
        cell.fill = PatternFill("solid", fgColor=SOURCE_COLORS["ranking"])
        cell.font = Font(color="FFFFFF", bold=True)
    components = [
        ("SQX", COMPONENT_WEIGHTS["sqx"] / 100),
        ("Backtest MT5", COMPONENT_WEIGHTS["backtest"] / 100),
        ("Live MT5", COMPONENT_WEIGHTS["live"] / 100),
        ("Riesgo", COMPONENT_WEIGHTS["risk"] / 100),
    ]
    for row, values in enumerate(components, start=5):
        sheet.cell(row, 1, values[0])
        sheet.cell(row, 2, values[1]).number_format = "0%"
    thresholds = [
        ("Cobertura mínima", 35),
        ("Score prometedora", 60),
        ("Cobertura prometedora", 50),
        ("Score prioridad alta", 75),
        ("Cobertura prioridad alta", 75),
    ]
    sheet["A11"] = "Umbral"
    sheet["B11"] = "Valor"
    for cell in sheet[11][:2]:
        cell.fill = PatternFill("solid", fgColor=SOURCE_COLORS["comparison"])
        cell.font = Font(color="FFFFFF", bold=True)
    for row, values in enumerate(thresholds, start=12):
        sheet.cell(row, 1, values[0])
        sheet.cell(row, 2, values[1]).number_format = "0.0"
    sheet["D4"] = "Fórmula"
    sheet["E4"] = "Definición"
    for cell in sheet[4][3:5]:
        cell.fill = PatternFill("solid", fgColor=SOURCE_COLORS["ranking"])
        cell.font = Font(color="FFFFFF", bold=True)
    formulas = [
        ("Cobertura", "Σ (peso del componente × fracción de datos disponibles)"),
        ("Calidad", "Σ (score × peso × disponibilidad) / Cobertura"),
        ("Score promesa", "Calidad × (0.5 + 0.5 × Cobertura / 100)"),
        ("Prioridad alta", "Score ≥ 75, cobertura ≥ 75 y riesgo distinto de rojo"),
        ("Prometedora", "Score ≥ 60, cobertura ≥ 50 y riesgo distinto de rojo"),
        ("Revisar/Pausar", "El control de riesgo live está en rojo"),
    ]
    for row, values in enumerate(formulas, start=5):
        sheet.cell(row, 4, values[0])
        sheet.cell(row, 5, values[1])
    dictionary = [
        ("Edge Score", "SQX", "0–100", "Mayor es mejor", "Pilares profitability, consistency, risk y entry."),
        ("EGT", "SQX", "Aprox. 0–10", "Mayor es mejor", "Alineación del P/L mensual con pendientes de mercado."),
        ("Decay", "SQX", "% IS→OOS", "Menor es mejor", "Deterioro positivo; valores negativos indican mejora OOS."),
        ("Concordancia", "SQX vs MT5", "0–100", "Mayor es mejor", "Cercanía de PF, DD%, frecuencia y win rate."),
        ("Retención PF", "Live vs OOS", "Ratio", "Cerca o sobre 1", "PF live dividido por PF OOS."),
        ("Utilización DD", "Live vs OOS", "Ratio", "Menor es mejor", "DD live dividido por límite OOS."),
        ("Racha actual", "Live", "Trades", "Menor es mejor", "Pérdidas consecutivas al cierre más reciente."),
        ("Cobertura", "Todas", "0–100", "Mayor es mejor", "No confundir falta de datos con mala calidad."),
    ]
    start = 20
    headers = ["KPI", "Fuente", "Unidad", "Lectura", "Definición"]
    indexes = _table(sheet, headers, [list(row) for row in dictionary], start, "MethodologyTable", SOURCE_COLORS["identity"])
    del indexes
    sheet["H4"] = "Cobertura del snapshot"
    sheet["H4"].fill = PatternFill("solid", fgColor=SOURCE_COLORS["identity"])
    sheet["H4"].font = Font(color="FFFFFF", bold=True)
    counts = [
        ("Estrategias", len(records)),
        ("Con MT5", sum(bool(record["mappings"]) for record in records)),
        ("Con SQX", sum(bool(record["strategy"].get("sqx")) for record in records)),
        ("Con Edge", sum(bool(record["analytics"].get("edge", {}).get("available")) for record in records)),
        ("Con EGT", sum(bool(record["analytics"].get("egt", {}).get("available")) for record in records)),
        ("Con backtest", sum(bool(record["completed"]) for record in records)),
    ]
    for row, values in enumerate(counts, start=5):
        sheet.cell(row, 8, values[0])
        sheet.cell(row, 9, values[1]).number_format = "#,##0"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 70
    sheet.column_dimensions["H"].width = 24
    sheet.column_dimensions["I"].width = 14
    for column in ("A", "B", "C", "D", "E"):
        sheet.column_dimensions[column].width = max(sheet.column_dimensions[column].width or 0, 18)
    sheet.freeze_panes = "A4"


def _comparative_sheet(
    workbook: Any,
    records: list[dict[str, Any]],
    generated_at: str,
    formula_cache: dict[str, dict[str, Any]],
) -> dict[int, int]:
    sheet = workbook["Comparativo"]
    headers = [
        "ID", "Estrategia", "Símbolo", "Estado vínculo", "Fuentes",
        "Score SQX", "Score backtest", "Score live", "Score riesgo",
        "Cobertura", "Calidad", "Score promesa", "Categoría", "Riesgo",
        "Edge", "EGT", "PF OOS", "PF backtest", "PF live",
        "Trades/mes OOS", "Trades/mes backtest", "Trades/mes live",
        "DD OOS", "DD % backtest", "DD live", "Racha OOS",
        "Racha backtest", "Racha live", "Racha actual",
        "Retención PF live/OOS", "Retención frecuencia", "Utilización DD",
        "Utilización racha", "Datos faltantes",
        "Disponibilidad SQX", "Disponibilidad backtest", "Disponibilidad live",
        "Disponibilidad riesgo",
    ]
    _title(sheet, "Comparativo | Todas las fuentes", f"Una fila por estrategia · {generated_at}", len(headers))
    group_row = 4
    groups = [
        (1, 5, "Identidad", SOURCE_COLORS["identity"]),
        (6, 9, "Componentes", SOURCE_COLORS["ranking"]),
        (10, 14, "Ranking y riesgo", SOURCE_COLORS["risk"]),
        (15, 29, "KPIs clave", SOURCE_COLORS["sqx"]),
        (30, 33, "Comparaciones", SOURCE_COLORS["comparison"]),
        (34, 38, "Cobertura y faltantes", SOURCE_COLORS["backtest"]),
    ]
    for start, end, label, color in groups:
        sheet.merge_cells(
            start_row=group_row, start_column=start, end_row=group_row, end_column=end
        )
        cell = sheet.cell(group_row, start, label)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        for column in range(start, end + 1):
            sheet.cell(group_row, column).fill = PatternFill("solid", fgColor=color)
    rows: list[list[Any]] = []
    for record in records:
        strategy = record["strategy"]
        edge = record["analytics"].get("edge", {})
        egt = record["analytics"].get("egt", {})
        bt = record["bt"]
        live = record["live"]
        oos = record["oos"]
        source_labels = [
            label
            for label, present in (
                ("MT5", bool(record["mappings"])),
                ("SQX", bool(strategy.get("sqx"))),
                ("BT", bool(record["completed"])),
                ("Excel", "excel" in str(strategy.get("origin") or "")),
            )
            if present
        ]
        rows.append(
            [
                strategy["id"], strategy["sqx_name"], strategy["symbol"],
                strategy.get("link_state"), "+".join(source_labels),
                record["scores"]["sqx"], record["scores"]["backtest"],
                record["scores"]["live"], record["scores"]["risk"],
                None, None, None, None, record["risk_status"],
                _number(edge, "score") if edge.get("available") else None,
                _number(egt, "total") if egt.get("available") else None,
                _number(oos, "ProfitFactor"), _number(bt, "profit_factor"),
                _number(live, "profit_factor"), _number(oos, "AvgTradesPerMonth"),
                _number(bt, "trades_per_month"), _number(live, "trades_per_month"),
                _number(oos, "Drawdown"), record["bt_parts"].get("dd_pct"),
                _number(live, "max_drawdown"), _number(oos, "MaxConsecLoss"),
                _number(bt, "max_consecutive_losses"),
                _number(live, "max_consecutive_losses"),
                _number(live, "current_consecutive_losses"),
                None, None, None, None, record["missing"],
                record["component_coverage"]["sqx"],
                record["component_coverage"]["backtest"],
                record["component_coverage"]["live"],
                record["component_coverage"]["risk"],
            ]
        )
    indexes = _table(sheet, headers, rows, 5, "ComparisonTable", SOURCE_COLORS["identity"])
    data_start = 6
    data_end = 5 + len(rows)
    weight_cells = {
        "sqx": "'Metodología'!$B$5",
        "backtest": "'Metodología'!$B$6",
        "live": "'Metodología'!$B$7",
        "risk": "'Metodología'!$B$8",
    }
    row_by_strategy: dict[int, int] = {}
    for offset, record in enumerate(records):
        row = data_start + offset
        row_by_strategy[int(record["strategy"]["id"])] = row
        component = {key: get_column_letter(indexes[key]) for key in ("Score SQX", "Score backtest", "Score live", "Score riesgo")}
        availability = {key: get_column_letter(indexes[key]) for key in ("Disponibilidad SQX", "Disponibilidad backtest", "Disponibilidad live", "Disponibilidad riesgo")}
        coverage_terms = [
            f"{availability['Disponibilidad SQX']}{row}*{weight_cells['sqx']}",
            f"{availability['Disponibilidad backtest']}{row}*{weight_cells['backtest']}",
            f"{availability['Disponibilidad live']}{row}*{weight_cells['live']}",
            f"{availability['Disponibilidad riesgo']}{row}*{weight_cells['risk']}",
        ]
        numerator = [
            f"{component['Score SQX']}{row}*{availability['Disponibilidad SQX']}{row}*{weight_cells['sqx']}",
            f"{component['Score backtest']}{row}*{availability['Disponibilidad backtest']}{row}*{weight_cells['backtest']}",
            f"{component['Score live']}{row}*{availability['Disponibilidad live']}{row}*{weight_cells['live']}",
            f"{component['Score riesgo']}{row}*{availability['Disponibilidad riesgo']}{row}*{weight_cells['risk']}",
        ]
        coverage_cell = f"{get_column_letter(indexes['Cobertura'])}{row}"
        quality_cell = f"{get_column_letter(indexes['Calidad'])}{row}"
        promise_cell = f"{get_column_letter(indexes['Score promesa'])}{row}"
        risk_cell = f"{get_column_letter(indexes['Riesgo'])}{row}"
        sheet[coverage_cell] = f"=100*({'+'.join(coverage_terms)})"
        _cache_formula(formula_cache, "Comparativo", sheet[coverage_cell], record["coverage"])
        sheet[quality_cell] = f'=IFERROR(({"+".join(numerator)})/({"+".join(coverage_terms)}),"")'
        _cache_formula(formula_cache, "Comparativo", sheet[quality_cell], record["quality"])
        sheet[promise_cell] = f'=IF({quality_cell}="","",{quality_cell}*(0.5+0.5*{coverage_cell}/100))'
        _cache_formula(formula_cache, "Comparativo", sheet[promise_cell], record["promise"])
        category_cell = sheet.cell(row, indexes["Categoría"])
        category_cell.value = (
            f'=IF({risk_cell}="red","Revisar/Pausar",'
            f'IF({coverage_cell}<\'Metodología\'!$B$12,"Datos insuficientes",'
            f'IF(AND({promise_cell}>=\'Metodología\'!$B$15,{coverage_cell}>=\'Metodología\'!$B$16),"Prioridad alta",'
            f'IF(AND({promise_cell}>=\'Metodología\'!$B$13,{coverage_cell}>=\'Metodología\'!$B$14),"Prometedora","Observar"))))'
        )
        _cache_formula(formula_cache, "Comparativo", category_cell, record["category"])
        for target, left, right in (
            ("Retención PF live/OOS", "PF live", "PF OOS"),
            ("Retención frecuencia", "Trades/mes live", "Trades/mes OOS"),
            ("Utilización DD", "DD live", "DD OOS"),
            ("Utilización racha", "Racha live", "Racha OOS"),
        ):
            left_cell = f"{get_column_letter(indexes[left])}{row}"
            right_cell = f"{get_column_letter(indexes[right])}{row}"
            formula_cell = sheet.cell(row, indexes[target])
            formula_cell.value = f'=IFERROR({left_cell}/{right_cell},"")'
            left_value = sheet[left_cell].value
            right_value = sheet[right_cell].value
            cached = (
                float(left_value) / float(right_value)
                if left_value not in (None, "") and right_value not in (None, "", 0)
                else None
            )
            _cache_formula(formula_cache, "Comparativo", formula_cell, cached)
    sheet.freeze_panes = "F6"
    _widths(
        sheet,
        headers,
        {
            "Estrategia": 42,
            "Estado vínculo": 18,
            "Fuentes": 18,
            "Categoría": 20,
            "Riesgo": 12,
            "Datos faltantes": 48,
        },
    )
    _format_columns(
        sheet, indexes, data_start, data_end,
        percent=[
            "Disponibilidad SQX", "Disponibilidad backtest", "Disponibilidad live",
            "Disponibilidad riesgo",
        ],
        money=["DD OOS", "DD live"],
        decimals=[
            "Score SQX", "Score backtest", "Score live", "Score riesgo", "Cobertura",
            "Calidad", "Score promesa", "Edge", "EGT", "PF OOS", "PF backtest",
            "PF live", "Trades/mes OOS", "Trades/mes backtest", "Trades/mes live",
            "DD % backtest",
        ],
        integers=["Racha OOS", "Racha backtest", "Racha live", "Racha actual"],
        ratios=["Retención PF live/OOS", "Retención frecuencia", "Utilización DD", "Utilización racha"],
    )
    for header in ("Disponibilidad SQX", "Disponibilidad backtest", "Disponibilidad live", "Disponibilidad riesgo"):
        sheet.column_dimensions[get_column_letter(indexes[header])].hidden = True
    for row in range(data_start, data_end + 1):
        sheet.cell(row, indexes["Datos faltantes"]).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        sheet.row_dimensions[row].height = 28
    if rows:
        for header in ("Score SQX", "Score backtest", "Score live", "Score riesgo", "Score promesa"):
            column = get_column_letter(indexes[header])
            sheet.conditional_formatting.add(
                f"{column}{data_start}:{column}{data_end}",
                ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                               mid_type="num", mid_value=60, mid_color="FFEB84",
                               end_type="num", end_value=100, end_color="63BE7B"),
            )
    return row_by_strategy


def _ranking_sheet(
    workbook: Any,
    records: list[dict[str, Any]],
    comparison_rows: dict[int, int],
    generated_at: str,
    formula_cache: dict[str, dict[str, Any]],
) -> None:
    sheet = workbook["Ranking"]
    headers = [
        "Rank", "Estrategia", "Símbolo", "Categoría", "Score promesa", "Cobertura",
        "Calidad", "Riesgo", "Vínculo", "Edge", "EGT", "PF OOS", "PF backtest",
        "PF live", "DD live", "Racha live", "Racha actual", "Trades live",
        "Datos faltantes",
    ]
    _title(sheet, "Ranking multifuente de estrategias", f"SQX + Backtest MT5 + Live broker · {generated_at}", len(headers))
    ordered = sorted(
        records,
        key=lambda record: (
            record["promise"] is not None,
            record["promise"] or -1,
            record["coverage"],
        ),
        reverse=True,
    )
    start_row = 10
    end_row = start_row + len(ordered) - 1
    cards = [
        ("A4:C4", "A5:C6", "Estrategias", f"=COUNTA(B{start_row}:B{end_row})"),
        ("E4:G4", "E5:G6", "Prioridad alta", f'=COUNTIF(D{start_row}:D{end_row},"Prioridad alta")'),
        ("I4:K4", "I5:K6", "Cobertura ≥ 75", f'=COUNTIF(F{start_row}:F{end_row},">=75")'),
        ("M4:O4", "M5:O6", "Riesgo rojo", f'=COUNTIF(H{start_row}:H{end_row},"red")'),
    ]
    for label_range, value_range, label, formula in cards:
        sheet.merge_cells(label_range)
        sheet[label_range.split(":")[0]] = label
        sheet[label_range.split(":")[0]].font = Font(color="52606D", bold=True, size=10)
        sheet.merge_cells(value_range)
        anchor = value_range.split(":")[0]
        sheet[anchor] = formula
        sheet[anchor].font = Font(color="102A43", bold=True, size=20)
        sheet[anchor].alignment = Alignment(vertical="center")
    card_values = [
        len(ordered),
        sum(record["category"] == "Prioridad alta" for record in ordered),
        sum(record["coverage"] >= 75 for record in ordered),
        sum(record["risk_status"] == "red" for record in ordered),
    ]
    for (_, value_range, _, _), value in zip(cards, card_values):
        _cache_formula(
            formula_cache,
            "Ranking",
            sheet[value_range.split(":")[0]],
            value,
        )
    rows = [
        [
            rank,
            record["strategy"]["sqx_name"],
            record["strategy"]["symbol"],
            None, None, None, None,
            record["risk_status"],
            record["strategy"].get("link_state"),
            _number(record["analytics"].get("edge", {}), "score")
            if record["analytics"].get("edge", {}).get("available") else None,
            _number(record["analytics"].get("egt", {}), "total")
            if record["analytics"].get("egt", {}).get("available") else None,
            _number(record["oos"], "ProfitFactor"),
            _number(record["bt"], "profit_factor"),
            _number(record["live"], "profit_factor"),
            _number(record["live"], "max_drawdown"),
            _number(record["live"], "max_consecutive_losses"),
            _number(record["live"], "current_consecutive_losses"),
            _number(record["live"], "trades"),
            record["missing"],
        ]
        for rank, record in enumerate(ordered, start=1)
    ]
    indexes = _table(sheet, headers, rows, 9, "RankingTable", SOURCE_COLORS["ranking"])
    comparison_columns = {
        "Categoría": "Categoría",
        "Score promesa": "Score promesa",
        "Cobertura": "Cobertura",
        "Calidad": "Calidad",
    }
    comparison_headers = [
        "ID", "Estrategia", "Símbolo", "Estado vínculo", "Fuentes",
        "Score SQX", "Score backtest", "Score live", "Score riesgo",
        "Cobertura", "Calidad", "Score promesa", "Categoría",
    ]
    comparison_index = {header: index + 1 for index, header in enumerate(comparison_headers)}
    for offset, record in enumerate(ordered):
        row = start_row + offset
        source_row = comparison_rows[int(record["strategy"]["id"])]
        for target, source in comparison_columns.items():
            formula_cell = sheet.cell(row, indexes[target])
            formula_cell.value = (
                f"='Comparativo'!{get_column_letter(comparison_index[source])}{source_row}"
            )
            cached_values = {
                "Categoría": record["category"],
                "Score promesa": record["promise"],
                "Cobertura": record["coverage"],
                "Calidad": record["quality"],
            }
            _cache_formula(
                formula_cache,
                "Ranking",
                formula_cell,
                cached_values[target],
            )
    sheet.freeze_panes = "D10"
    _widths(
        sheet,
        headers,
        {
            "Estrategia": 42,
            "Categoría": 20,
            "Riesgo": 12,
            "Vínculo": 18,
            "Datos faltantes": 52,
        },
    )
    _format_columns(
        sheet, indexes, start_row, end_row,
        money=["DD live"],
        decimals=[
            "Score promesa", "Cobertura", "Calidad", "Edge", "EGT",
            "PF OOS", "PF backtest", "PF live",
        ],
        integers=["Rank", "Racha live", "Racha actual", "Trades live"],
    )
    for row in range(start_row, end_row + 1):
        sheet.cell(row, indexes["Datos faltantes"]).alignment = Alignment(
            vertical="top", wrap_text=True
        )
        sheet.row_dimensions[row].height = 28
    if ordered:
        score_column = get_column_letter(indexes["Score promesa"])
        coverage_column = get_column_letter(indexes["Cobertura"])
        sheet.conditional_formatting.add(
            f"{score_column}{start_row}:{score_column}{end_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="0F766E"),
        )
        sheet.conditional_formatting.add(
            f"{coverage_column}{start_row}:{coverage_column}{end_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="2563EB"),
        )
        category_column = get_column_letter(indexes["Categoría"])
        category_fills = {
            "Prioridad alta": "C6EFCE",
            "Prometedora": "DDEBF7",
            "Observar": "FFF2CC",
            "Datos insuficientes": "E7E6E6",
            "Revisar/Pausar": "FFC7CE",
        }
        for category, color in category_fills.items():
            sheet.conditional_formatting.add(
                f"{category_column}{start_row}:{category_column}{end_row}",
                FormulaRule(
                    formula=[f'${category_column}{start_row}="{category}"'],
                    fill=PatternFill("solid", fgColor=color),
                ),
            )
        chart_rows = min(15, len(ordered))
        if chart_rows:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "Top 15 por Score promesa"
            chart.x_axis.title = "Score (0-100)"
            chart.height = 8
            chart.width = 15
            data = Reference(
                sheet,
                min_col=indexes["Score promesa"],
                min_row=9,
                max_row=9 + chart_rows,
            )
            categories = Reference(
                sheet,
                min_col=indexes["Estrategia"],
                min_row=10,
                max_row=9 + chart_rows,
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.legend = None
            sheet.add_chart(chart, "U4")


def export_catalog(
    source: Path,
    destination: Path,
    strategies: list[dict[str, Any]],
) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if source.is_file():
        workbook = load_workbook(source)
    else:
        workbook = Workbook()
        workbook.active.title = "SQX_strategy list"
    for name in GENERATED_SHEETS:
        if name in workbook.sheetnames:
            del workbook[name]
    for index, name in enumerate(GENERATED_SHEETS[:-1]):
        workbook.create_sheet(name, index)
    records = _build_records(strategies)
    formula_cache: dict[str, dict[str, Any]] = {}
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    _methodology_sheet(workbook, records, generated_at)
    comparison_rows = _comparative_sheet(
        workbook,
        records,
        generated_at,
        formula_cache,
    )
    _sqx_sheet(workbook, records, generated_at)
    _backtest_sheet(workbook, records, generated_at)
    _live_sheet(workbook, records, generated_at)
    _ranking_sheet(
        workbook,
        records,
        comparison_rows,
        generated_at,
        formula_cache,
    )
    workbook.active = workbook.sheetnames.index("Ranking")
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.modified = datetime.now()
    workbook.save(destination)
    _write_formula_cache(destination, formula_cache)
    return destination
